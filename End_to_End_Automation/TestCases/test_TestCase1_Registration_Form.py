from selenium.webdriver.common.by import By
import pytest
import openpyxl
from Selenium_Automation.End_to_End_Automation.Base.base import open_browser, close_browser
from Selenium_Automation.End_to_End_Automation.Library import ConfigReader
from Selenium_Automation.End_to_End_Automation.Pages import RegistrationPage

def data_generator():
    data = []
    wb = openpyxl.load_workbook('/Users/ivanzhelev/Documents/PersonalProjects/Selenium_Automation/End_to_End_Automation/Book1.xlsx')
    sh = wb['Sheet1']
    rows = sh.max_row
    for r in range(1, rows + 1):
        username = sh.cell(r, 1).value
        password = sh.cell(r, 2).value
        data.append([username, password])

    return data


@pytest.mark.parametrize('data', data_generator())
def test_Validate_Registration(data):

    browser = open_browser()

    registration = RegistrationPage.RegistrationPage(browser)
    registration.enter_username(data[0])
    # registration.enter_email('ivan@ivan.com')
    registration.enter_password(data[1])
    registration.enter_confirm_password(data[1])

    close_browser()

